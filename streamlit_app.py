import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from pydantic import BaseModel
import openpyxl
from openpyxl import load_workbook
import tempfile
import base64  # Ensure base64 is imported
import json
import os
import openai

# Classes and Functions
class Purchase(BaseModel):
    alcohol: str
    date: str
    store: str
    quantity: int
    price: int

class Bill(BaseModel):
    purchases: list[Purchase]

def append_df_to_excel(df, excel_file_path, sheet_name='Sheet1'):
    """
    Append a DataFrame to an existing Excel file, starting from the first truly empty row.

    Parameters:
    - df: DataFrame to append
    - excel_file_path: Path to the existing Excel file
    - sheet_name: The sheet name where the DataFrame will be appended. Default is 'Sheet1'.
    """
    # Load the existing workbook
    book = load_workbook(excel_file_path)
    
    # Access the sheet, or create it if it does not exist
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        
        # Find the first empty row by checking cell contents
        startrow = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
            if all(cell.value is None for cell in row):
                startrow = row[0].row - 1
                break
        else:
            # If no empty row found, startrow is the max_row
            startrow = sheet.max_row
    else:
        startrow = 0  # Start from the beginning if the sheet does not exist

    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, 
                    startrow=startrow, startcol=1)

def authenticate_gsheet():
    # Load credentials from the downloaded JSON file
    credentials = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    client = gspread.authorize(credentials)
    return client

def encode_image(image_path):
    """Encodes an image to base64 format."""
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        st.error(f"Error encoding image: {e}")
        return None

def parse_purchases_to_dataframe(purchases_json):
    # Load the JSON string into a Python dictionary
    purchases_dict = json.loads(purchases_json)
    
    # Convert the 'purchases' list from the dictionary into a DataFrame
    df = pd.DataFrame(purchases_dict['purchases'])
    
    return df

def check_password():
    """Returns `True` if the user has the correct password."""
    def password_entered():
        if st.session_state["password"] == st.secrets["app"]["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.text_input("Password", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.text_input("Password", type="password", on_change=password_entered, key="password")
        st.error("😕 Password incorrect")
        return False
    else:
        # Password correct.
        return True

# Streamlit Logic
def streamlit_app():
    st.title('Bill Uploader')

    if check_password():
        openai.api_key = st.secrets["openai"]["api_key"]
        st.write("Password Accepted")
        # Option to choose storage method
        storage_option = st.radio("Select storage option", ('Excel File', 'Google Sheets'))

        if storage_option == 'Excel File':
            # File uploader for selecting an Excel file
            excel_file = st.file_uploader("Upload an Excel file to modify", type='xlsx')
            if excel_file:
                # Use tempfile for handling the Excel file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(excel_file.getbuffer())
                    tmp_path = tmp.name

                uploaded_files = st.file_uploader("Choose images...", type='jpg', accept_multiple_files=True)

                if uploaded_files is not None:
                    all_data = []

                    for uploaded_file in uploaded_files:
                        # Use tempfile for handling image files temporarily
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_img:
                            temp_img.write(uploaded_file.read())
                            temp_file_path = temp_img.name

                        encoded_image = encode_image(temp_file_path)

                        if encoded_image is None:
                            continue  # Skip if there was an error encoding the image

                        try:
                            response = openai.beta.chat.completions.parse(
                                model="gpt-4o-mini",  
                                messages=[
                                    {
                                        'role': 'user',
                                        'content': [
                                            {'type': 'text', 'text': """You are an expert at structured data extraction. From the picture of this bill, get the Alcohol Name, Date Purchased (MM/DD/YYYY), Store Name, Quantity Purchased, and Price per Bottle. Output the data into the given structure."""},
                                            {'type': 'image_url', 'image_url': {'url': f'data:image/jpeg;base64,{encoded_image}'}}
                                        ]
                                    }
                                ],
                                response_format=Bill
                            )
                            result = response.choices[0].message.content
                            frame = parse_purchases_to_dataframe(result)
                            all_data.append(frame)

                            # Debugging: Print the dataframe for each processed image
                            st.write(f"Extracted data for {uploaded_file.name}:")
                            st.dataframe(frame)

                        except Exception as e:
                            st.error(f"Error processing bill text from file {uploaded_file.name}: {e}")
                        finally:
                            # Remove the temporary file after processing
                            os.remove(temp_file_path)

                    # Combine all extracted data into a single DataFrame
                    if all_data:
                        combined_frame = pd.concat(all_data, ignore_index=True)
                        
                        st.write("Here is the extracted data from all uploaded files:")
                        st.dataframe(combined_frame)
                        
                        # User confirmation to append data to Excel
                        if st.button("Is the data correct? Click to append to Excel."):
                            append_df_to_excel(combined_frame, tmp_path)
                            st.success("Data appended to Excel successfully!")

                            # Provide download link
                            with open(tmp_path, "rb") as f:
                                excel_data = f.read()
                            st.download_button(label="Download modified Excel file", data=excel_data, file_name="modified_excel_file.xlsx")

        elif storage_option == 'Google Sheets':
            # Google Sheets authentication
            client = authenticate_gsheet()

            # User inputs for Google Sheet ID and range
            sheet_id = st.text_input("Enter your Google Sheet ID")
            sheet_name = st.text_input("Enter the sheet name (e.g., Sheet1)")

            if sheet_id and sheet_name:
                uploaded_files = st.file_uploader("Choose images...", type='jpg', accept_multiple_files=True)

                if uploaded_files is not None:
                    all_data = []

                    for uploaded_file in uploaded_files:
                        # Use tempfile for handling image files temporarily
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_img:
                            temp_img.write(uploaded_file.read())
                            temp_file_path = temp_img.name

                        encoded_image = encode_image(temp_file_path)

                        if encoded_image is None:
                            continue  # Skip if there was an error encoding the image

                        try:
                            response = openai.beta.chat.completions.parse(
                                model="gpt-4o-mini",  
                                messages=[
                                    {
                                        'role': 'user',
                                        'content': [
                                            {'type': 'text', 'text': """You are an expert at structured data extraction. From the picture of this bill, get the Alcohol Name, Date Purchased (MM/DD/YYYY), Store Name, Quantity Purchased, and Price per Bottle. Output the data into the given structure."""},
                                            {'type': 'image_url', 'image_url': {'url': f'data:image/jpeg;base64,{encoded_image}'}}
                                        ]
                                    }
                                ],
                                response_format=Bill
                            )
                            result = response.choices[0].message.content
                            frame = parse_purchases_to_dataframe(result)
                            all_data.append(frame)

                            # Debugging: Print the dataframe for each processed image
                            st.write(f"Extracted data for {uploaded_file.name}:")
                            st.dataframe(frame)

                        except Exception as e:
                            st.error(f"Error processing bill text from file {uploaded_file.name}: {e}")
                        finally:
                            # Remove the temporary file after processing
                            os.remove(temp_file_path)

                    # Combine all extracted data into a single DataFrame
                    if all_data:
                        combined_frame = pd.concat(all_data, ignore_index=True)
                        
                        st.write("Here is the extracted data from all uploaded files:")
                        st.dataframe(combined_frame)

                        # User confirmation to append data to Google Sheets
                        if st.button("Is the data correct? Click to append to Google Sheets."):
                            try:
                               # Open the Google Sheet by ID
                                sheet = client.open_by_key(sheet_id)
                                worksheet = sheet.worksheet(sheet_name)
                                
                                # Get all data from the worksheet (if needed to display or manipulate)
                                existing_data = worksheet.get_all_values()
                                existing_df = pd.DataFrame(existing_data[1:], columns=existing_data[0])
                                
                                # Display existing data (if required)
                                st.write("Existing data in Google Sheet:")
                                st.dataframe(existing_df)
                                
                                # Convert combined_frame to list format for appending
                                combined_data = combined_frame.values.tolist()
                                
                                # Append data to the worksheet
                                worksheet.append_rows(combined_data)
                                
                                st.success("Data appended to Google Sheet successfully!")
                                # Update Google Sheet with new data
                                worksheet.update([combined_df.columns.values.tolist()] + combined_df.values.tolist())
                                st.success("Google Sheet updated successfully!")
                            except Exception as e:
                                st.error(f"An error occurred: {e}")

if __name__ == "__main__":
    streamlit_app()
